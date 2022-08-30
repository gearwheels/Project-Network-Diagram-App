
import sys, os
import numpy
from pathlib import Path
### Для обработки .xlsx файлов ##############
import openpyxl

### Для обработки .pdf файлов ###############

from borb.pdf import Document
from borb.pdf import Page
from borb.pdf import SingleColumnLayout
from borb.pdf import Paragraph
from borb.pdf import PDF



from PyQt5 import QtWidgets
from PyQt5.QtCore import QRect
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QAction

############# Кастомные файлы для проги #####################
from MainMenu import Ui_MainMenu
from windowTask1 import Ui_MainWindow1
from windowTask3 import Ui_MainWindow3
from windowTask2 import Ui_MainWindow2
from tableTask2 import Ui_tableTask2Widget
from windowTask6 import Ui_MainWindow6
import Display
import WinsDialog
from WinsDialog import winSigReport,winLogin,winEditTable
from Color import Color
from task1CheckForm import task1CheckForm
import graph_model as gm
import EditTable


#////////////////////////////////  КЛАСС ОКНА ПЕРВОГО ЗАДАНИЯ  ////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class Window1(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)

        self.ui = Ui_MainWindow1()
        self.ui.setupUi(self)

        self.setWindowTitle("Задача №1")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        

        self.centralWidget = Display.Display(self)
        self.setCentralWidget(self.centralWidget)

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()


    def addNode(self):
        if self.ui.actionbtnAddNode.isChecked() == False:
            self.centralWidget.functionAble = ""
        else:
            self.centralWidget.functionAble = "Добавить вершину"
            self.ui.actionbtnConnectNode.setChecked(False)
            self.ui.actionbtnRemoveNodeConnection.setChecked(False)
            self.ui.actionbtnMoveNode.setChecked(False)
            self.ui.actionbtnRemoveNode.setChecked(False)

    def addArrow(self):
        if self.ui.actionbtnConnectNode.isChecked() == False:
            self.centralWidget.functionAble = ""
        else:
            self.centralWidget.functionAble = "Добавить связь"
            self.ui.actionbtnAddNode.setChecked(False)
            self.ui.actionbtnRemoveNodeConnection.setChecked(False)
            self.ui.actionbtnMoveNode.setChecked(False)
            self.ui.actionbtnRemoveNode.setChecked(False)

    def removeArrow(self):
        if self.ui.actionbtnRemoveNodeConnection.isChecked() == False:
            self.centralWidget.functionAble = ""
        else:
            self.centralWidget.functionAble = "Удалить связь"
            self.ui.actionbtnConnectNode.setChecked(False)
            self.ui.actionbtnAddNode.setChecked(False)
            self.ui.actionbtnMoveNode.setChecked(False)
            self.ui.actionbtnRemoveNode.setChecked(False)

    def removeNode(self):
        if self.ui.actionbtnRemoveNode.isChecked() == False:
            self.centralWidget.functionAble = ""
        else:
            self.centralWidget.functionAble = "Удалить вершину"
            self.ui.actionbtnConnectNode.setChecked(False)
            self.ui.actionbtnAddNode.setChecked(False)
            self.ui.actionbtnMoveNode.setChecked(False)
            self.ui.actionbtnRemoveNodeConnection.setChecked(False)


    def moveNode(self):
        if self.ui.actionbtnMoveNode.isChecked() == False:
            self.centralWidget.functionAble = ""
        else:
            self.centralWidget.functionAble = "Переместить вершины"
            self.ui.actionbtnConnectNode.setChecked(False)
            self.ui.actionbtnAddNode.setChecked(False)
            self.ui.actionbtnRemoveNodeConnection.setChecked(False)
            self.ui.actionbtnRemoveNode.setChecked(False)

    def makeNewFile(self):
        self.centralWidget.functionAble = "Новый файл"

    def sizeGet(self):
        return self.size()

    def taskCheck(self):
        mistakes = self.centralWidget.checkEvent()
        self.checkForm1 = task1CheckForm(self, mistakes)
        self.checkForm1.exec_()

    def _connectAction(self):
        self.ui.actionbtnAddNode.triggered.connect(self.addNode)
        self.ui.actionbtnConnectNode.triggered.connect(self.addArrow)
        self.ui.actionbtnRemoveNodeConnection.triggered.connect(self.removeArrow) # названия actionbtnRemoveNodeConnection и actionbtnRemoveNode надо поменять местами или иконки поменять местами
        self.ui.actionbtnMoveNode.triggered.connect(self.moveNode)
        self.ui.actionbtnRemoveNode.triggered.connect(self.removeNode)
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)
        self.ui.actionbtnCheck.triggered.connect(self.taskCheck)

    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def show(self):
        self.showMaximized()


#////////////////////////////////  КЛАСС ОКНА ВТОРОГО ЗАДАНИЯ  ////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class Window2(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Создаём компоновщик
        self.layout = QtWidgets.QHBoxLayout()
        # Добавляем виджет отрисовки в компоновщик
        self.DisplayObj = Display.Display2(self)
        self.layout.addWidget(self.DisplayObj)
        # Создаём виджет таблицы и добавляем его в компоновщик
        self.layout2 = QtWidgets.QVBoxLayout()
        self.table1 = QWidget()
        self.table1.ui = Ui_tableTask2Widget()
        self.table1.ui.setupUi(self.table1)
        self.table2 = QWidget()
        self.table2.ui = Ui_tableTask2Widget()
        self.table2.ui.setupUi(self.table2)
        self.table2.ui.tableWidget.setHorizontalHeaderLabels(["Поздний срок"])
        self.layout2.addWidget(self.table1)
        self.layout2.addWidget(self.table2)
        self.widget2 = QWidget()
        self.widget2.setLayout(self.layout2)
        
        self.layout.addWidget(self.widget2)
        # Задаём растяжение объектов в компоновщике
        self.layout.setStretch(0, 1)
        # Задаём компоновку виджету
        self.widget = QWidget()
        self.widget.setLayout(self.layout)

        self.ui = Ui_MainWindow2()
        self.ui.setupUi(self)
        # Присваиваем виджет с компоновкой окну
        self.setCentralWidget(self.widget)

        self.setWindowTitle("Задача №2")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        width = int(sizeWindow.width() - sizeWindow.width() / 5 + 280)
        height = int(sizeWindow.height() - sizeWindow.height() / 5)
        # вписываем во весь экран
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 12), int(sizeWindow.height() / 12))

        # Создаём окно для ошибки заполнения таблицы
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Ошибка!")
        self.msg.setText("Заполните все поля таблицы!")
        self.msg.setIcon(QMessageBox.Critical)
        self.msg.setStandardButtons(QMessageBox.Ok)

        # self.checkForm = task1CheckForm(self) # диалоговое окно для проврки задания

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()

    def show(self):
        # При вызове окна обновляется кол-во вершин графа
        self.showMaximized()
        self.cnt = len(Display.graph.Points)
        self.table1.ui.tableWidget.setRowCount(self.cnt)
        self.table2.ui.tableWidget.setRowCount(self.cnt)


    def table1Check(self):
        # Обнуляем данные в модели
        Display.graph.tp = numpy.empty((0))
        # Считываем новые
        for row in range(self.table1.ui.tableWidget.rowCount()):
            # Проверка на пустую ячейку
            if type(self.table1.ui.tableWidget.item(row, 0)) == QtWidgets.QTableWidgetItem and self.table1.ui.tableWidget.item(row, 0).text() != '': 
                # Добавление значения
                Display.graph.tp = numpy.append(Display.graph.tp, int(self.table1.ui.tableWidget.item(row, 0).text()))
            else:
                # При ошибке вызываем окно
                self.msg.show()
                break
        # print (Display.graph.tp)
        self.update()

    def table2Check(self):
        # То же самое для второй таблицы
        Display.graph.tn = numpy.empty((0))
        for row in range(self.table2.ui.tableWidget.rowCount()):
            if type(self.table2.ui.tableWidget.item(row, 0)) == QtWidgets.QTableWidgetItem and self.table2.ui.tableWidget.item(row, 0).text() != '':
                Display.graph.tn = numpy.append(Display.graph.tn, int(self.table2.ui.tableWidget.item(row, 0).text()))
            else:
                self.msg.show()
                break
        # print (Display.graph.tn)
        self.update()

    def critPathCheck(self):
        if self.ui.actionbtnCritPath.isChecked() == False:
            self.DisplayObj.functionAble = ""
        else:
            self.DisplayObj.functionAble = "Критический путь"

    def taskCheck(self):
        pass

    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def _connectAction(self):
        self.table1.ui.tableCheckButton.clicked.connect(self.table1Check)
        self.table2.ui.tableCheckButton.clicked.connect(self.table2Check)
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)
        self.ui.actionbtnCritPath.triggered.connect(self.critPathCheck)

    def sizeGet(self):
        return self.size()


#////////////////////////////////  КЛАСС ОКНА ТРЕТЬЕГО ЗАДАНИЯ  ///////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class Window3(QMainWindow):

    """ def __init__(self, parent=None):
        super().__init__(parent)

        # Создаём компоновщик
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(Color('blue'))
        layout.addWidget(Color('red'))
        # Задаём компоновку виджету
        widget = QWidget()
        widget.setLayout(layout)

        self.ui = Ui_MainWindow2()
        self.ui.setupUi(self)
        # Присваиваем виджет с компоновкой окну
        self.setCentralWidget(widget)

        self.setWindowTitle("Задача №3")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        width = int(sizeWindow.width() - sizeWindow.width() / 5)
        height = int(sizeWindow.height() - sizeWindow.height() / 5)
        # вписываем во весь экран
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 10), int(sizeWindow.height() / 10))

        # self.centralWidget = Display()
        # self.setCentralWidget(self.centralWidget)
        # self._connectAction() """

    def __init__(self, parent=None):
        super().__init__(parent)

        self.ui = Ui_MainWindow3()
        self.ui.setupUi(self)

        self.setWindowTitle("Задача №3")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        
        graph3 = gm.Graph(30)

        self.centralWidget = Display.Display3(self, 0, 0, 75, [0, 0, 255, 200], False, graph3)
        self.setCentralWidget(self.centralWidget)

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()

    def addNode(self):
        self.centralWidget.functionAble = "Добавить вершину"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def addArrow(self):
        self.centralWidget.functionAble = "Добавить связь"
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def addDottedArrow(self):
        self.centralWidget.functionAble = "Добавить пунктирную связь"
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def removeArrow(self):
        self.centralWidget.functionAble = "Удалить связь"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def removeNode(self):
        self.centralWidget.functionAble = "Удалить вершину"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)

    def moveNode(self):
        self.centralWidget.functionAble = "Переместить вершины"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def makeNewFile(self):
        self.centralWidget.functionAble = "Новый файл"

    def taskCheck(self):
        mistakes = self.centralWidget.checkEvent()
        self.checkForm1 = task1CheckForm(self, mistakes)
        self.checkForm1.exec_()

    def _connectAction(self):
        self.ui.actionbtnAddNode.triggered.connect(self.addNode)
        self.ui.actionbtnConnectNode.triggered.connect(self.addArrow)
        self.ui.actionbtnRemoveNodeConnection.triggered.connect(self.removeArrow)
        self.ui.actionbtnMoveNode.triggered.connect(self.moveNode)
        self.ui.actionbtnRemoveNode.triggered.connect(self.removeNode)
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)
        self.ui.actionbtnCheck.triggered.connect(self.taskCheck)
        self.ui.actionbtnDottedConnectNode.triggered.connect(self.addDottedArrow)

    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def show(self):
        self.showMaximized()

    def sizeGet(self):
        return self.size()

    


#////////////////////////////////  КЛАСС ОКНА ЧЕТВЁРТОГО ЗАДАНИЯ  /////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class Window4(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)

        self.ui = Ui_MainWindow3()
        self.ui.setupUi(self)

        self.setWindowTitle("Задача №4")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        

        self.centralWidget = Display.Display(self, 0, 0, 75, [0, 0, 255, 200], False)
        self.setCentralWidget(self.centralWidget)

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()

    def addNode(self):
        self.centralWidget.functionAble = "Добавить вершину"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def addArrow(self):
        self.centralWidget.functionAble = "Добавить связь"
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def addDottedArrow(self):
        self.centralWidget.functionAble = "Добавить пунктирную связь"
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def removeArrow(self):
        self.centralWidget.functionAble = "Удалить связь"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def removeNode(self):
        self.centralWidget.functionAble = "Удалить вершину"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnMoveNode.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)

    def moveNode(self):
        self.centralWidget.functionAble = "Переместить вершины"
        self.ui.actionbtnConnectNode.setChecked(False)
        self.ui.actionbtnAddNode.setChecked(False)
        self.ui.actionbtnRemoveNodeConnection.setChecked(False)
        self.ui.actionbtnDottedConnectNode.setChecked(False)
        self.ui.actionbtnRemoveNode.setChecked(False)

    def makeNewFile(self):
        self.centralWidget.functionAble = "Новый файл"

    def taskCheck(self):
        mistakes = self.centralWidget.checkEvent()
        self.checkForm1 = task1CheckForm(self, mistakes)
        self.checkForm1.exec_()

    def _connectAction(self):
        self.ui.actionbtnAddNode.triggered.connect(self.addNode)
        self.ui.actionbtnConnectNode.triggered.connect(self.addArrow)
        self.ui.actionbtnRemoveNodeConnection.triggered.connect(self.removeArrow)
        self.ui.actionbtnMoveNode.triggered.connect(self.moveNode)
        self.ui.actionbtnRemoveNode.triggered.connect(self.removeNode)
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)
        self.ui.actionbtnCheck.triggered.connect(self.taskCheck)
        self.ui.actionbtnDottedConnectNode.triggered.connect(self.addDottedArrow)

    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def show(self):
        self.showMaximized()

    def sizeGet(self):
        return self.size()


#////////////////////////////////  КЛАСС ОКНА ПЯТОЕ ЗАДАНИЯ  ////////////////////////////////////
#////////////////////////////////////////////////////////////////////////////////////////////////
class Window5(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)

        # Создаём компоновщик
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(Color('blue'))
        layout.addWidget(Color('red'))
        # Задаём компоновку виджету
        widget = QWidget()
        widget.setLayout(layout)

        self.ui = Ui_MainWindow3()
        self.ui.setupUi(self)
        # Присваиваем виджет с компоновкой окну
        self.setCentralWidget(widget)

        self.setWindowTitle("Задача №5")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        width = int(sizeWindow.width() - sizeWindow.width() / 5)
        height = int(sizeWindow.height() - sizeWindow.height() / 5)
        # вписываем во весь экран
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 10), int(sizeWindow.height() / 10))

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()

    def _connectAction(self):
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)


    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def show(self):
        self.showMaximized()


#////////////////////////////////  КЛАСС ОКНА ШЕСТОГО ЗАДАНИЯ  ////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class Window6(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)

        # Создаём компоновщик
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(Color('white'))
        layout.addWidget(Color('blue'))
        layout.addWidget(Color('red'))
        # Задаём компоновку виджету
        widget = QWidget()
        widget.setLayout(layout)

        self.ui = Ui_MainWindow6()
        self.ui.setupUi(self)
        # Присваиваем виджет с компоновкой окну
        self.setCentralWidget(widget)

        self.setWindowTitle("Задача №6")
        sizeWindow = QRect(QApplication.desktop().screenGeometry())
        width = int(sizeWindow.width() - sizeWindow.width() / 5)
        height = int(sizeWindow.height() - sizeWindow.height() / 5)
        # вписываем во весь экран
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 10), int(sizeWindow.height() / 10))

        self._connectAction()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

    def closeEvent(self, event):
        if self.ui.actionbtnHome.isChecked():
            self.ui.actionbtnHome.setChecked(False)
            event.accept()
        else:
            close = QMessageBox()
            close.setText("Вы уверены,что хотите закрыть программу?")
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Ok:
                event.accept()
            else:
                event.ignore()


    def _connectAction(self):
        self.ui.actionbtnHome.triggered.connect(self.backMainMenu)

    def backMainMenu(self):
        MainWindow.show()
        self.close()

    def show(self):
        self.showMaximized()


#////////////////////////////////////  КЛАСС ОКНА МЕНЮ  ///////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
class WindowMenu(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)

        self.ui = Ui_MainMenu()
        self.ui.setupUi(self)

        self.setWindowTitle("Меню")
        self.sizeWindow = QRect(QApplication.desktop().screenGeometry())

        self.name = "Иван"      # данные о студенте проинициализированы
        self.surname = "Иванов" # данные о студенте проинициализированы
        self.numGroup = "1"   # данные о студенте проинициализированы
        self.numINGroup = "9"  # данные о студенте проинициализированы

        self.startWindow = winLogin(self)# стартовое диалоговое окно для подписти отчета (имя фамилия номер группы)
        self.startWindow.exec_() # его запуск в отдельном потоке
        self.winSigReport = winSigReport(self) # диалоговое окно для подписти отчета (имя фамилия номер группы)
        self.winEditTable = winEditTable(self) #
        #self.creatTable = WinsDialog.creatTable(self) #

        self._connectAction()
        #self.creatReport()

        quit = QAction("Quit", self)
        quit.triggered.connect(self.closeEvent)

        self.testGen()

    def closeEvent(self, event):
        close = QMessageBox()
        close.setText("Вы уверены,что хотите закрыть программу?")
        close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        close = close.exec()

        if close == QMessageBox.Ok:
            event.accept()
        else:
            event.ignore()

    def _connectAction(self):
        self.ui.btnTask1.clicked.connect(lambda: self.openTask(self.ui.btnTask1.text()))
        self.ui.btnTask2.clicked.connect(lambda: self.openTask(self.ui.btnTask2.text()))
        self.ui.btnTask3.clicked.connect(lambda: self.openTask(self.ui.btnTask3.text()))
        self.ui.btnTask4.clicked.connect(lambda: self.openTask(self.ui.btnTask4.text()))
        self.ui.btnTask5.clicked.connect(lambda: self.openTask(self.ui.btnTask5.text()))
        self.ui.btnTask6.clicked.connect(lambda: self.openTask(self.ui.btnTask6.text()))
        self.ui.btnReportSign.clicked.connect(self.winSigReport.exec) # по клику вызываем диалоговое окно для подписти отчета и передаем управление ему
        self.ui.btnGenVar.clicked.connect(lambda: self.testGen()) # по клику генерируем задание (заполняем таблицу)
        #self.ui.previewReport.clicked.connect(lambda: self.creatReport()) #
        self.ui.btnEditTaskVariant.clicked.connect(self.winEditTable.exec)

    def activateDeveloperMode(self):
        self.name = "Иван"  # данные о студенте проинициализированы
        self.surname = "Иванов"  # данные о студенте проинициализированы
        self.numGroup = "1"  # данные о студенте проинициализированы
        self.numINGroup = "9"  # данные о студенте проинициализированы

    #def creatReport(self):
        # create an empty Document
        #pdf = Document()

        # add an empty Page
        #page = Page()
        #pdf.add_page(page)

        # use a PageLayout (SingleColumnLayout in this case)
        #layout = SingleColumnLayout(page)

        # add a Paragraph object
        #layout.add(Paragraph(self.name))
        #layout.add(Paragraph(self.surname))
        #layout.add(Paragraph(self.numGroup))
        #layout.add(Paragraph(self.numINGroup))

        # store the PDF
        #with open(Path("output.pdf"), "wb") as pdf_file_handle:
        #    PDF.dumps(pdf_file_handle, pdf)



    def openTask (self, numTask):
        if numTask == "Задание 1":
            MainWindow1.show()
        elif numTask == "Задание 2":
            MainWindow2.show()
        elif numTask == "Задание 3":
            MainWindow3.show()
        elif numTask == "Задание 4":
            MainWindow4.show()
        elif numTask == "Задание 5":
            MainWindow5.show()
        elif numTask == "Задание 6":
            MainWindow6.show()
        self.hide()

    def show(self):
        self.showMaximized()
        self.ui.tableVar.horizontalHeader().setDefaultSectionSize(int(self.sizeWindow.width() / 5))

    def testGen(self):  # функция записи в таблицу лабы конкретного задания (цифр: номер работы, номер отделения, кол-во часов и тд)

        fileName = "В" + self.numINGroup + ".xlsx" # выбираем нужную табличку по названию
        # файлик с таблицой должен называться "В" + номер студента по списку + ".xlsx" (расширение файла)
        pathFileXlsx = os.path.join("resources", "variants", fileName)# находим путь до файла
        book = openpyxl.open(pathFileXlsx, read_only=True) # открываем файл с помощью либы для обработки .xlsx
        sheet = book.active # active - выбирает номер страницы в книге без параметров (по умолчанию) первая страница

        # check = sheet.max_row
        # for i in range(1, sheet.max_row + 1):  # Начинаем c 1 строки, так как если начнем с 0 вылетит ошибка
        #    for j in range(0, sheet.max_column):  # Проходимся по кол-ву столбцов
        #        print(sheet[sheet[ i ][ j ].coordinate].value, end=" ")  # Выводим в консоль каждый элемент в каждой строке
        #    print()  # Перенос на новую строку

        countColumns = 0
        tabelVar = []

        for row in sheet.iter_rows(sheet.min_row, sheet.max_row):# подкачиваем данные из xlsx файла
            rowVar = []
            for cell in row:
                rowVar.append(cell.value)
            tabelVar.append(rowVar)

        #for list in tabelVar:
        #    print(list)

        self.ui.tableVar.setRowCount(0) # удаление старых данных из таблицы (если уже генерировалась таблица с заданием)

        for list in tabelVar:
            rowPosition = self.ui.tableVar.rowCount()  # генерируем строку в таблице для записи в нее чиселок
            self.ui.tableVar.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"
            for item in list:
                if countColumns >= 0:
                    # print(item, end=" ")
                    self.ui.tableVar.setItem(rowPosition, countColumns, QtWidgets.QTableWidgetItem(item))  # заполняем "строку таблицы из файла", каждую ячейку
                countColumns = countColumns + 1
            countColumns = 0
            # print()


        # for row in sheet.iter_rows(sheet.min_row+1, sheet.max_row):
        #    rowPosition = self.ui.tableVar.rowCount()  # генерируем строку в таблице для записи в нее чиселок
        #    self.ui.tableVar.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"

        #    for cell in row:
        #        if countColumns > 0:
        #            print(cell.value, end=" ")
        #            self.ui.tableVar.setItem(rowPosition, countColumns - 1, QtWidgets.QTableWidgetItem(cell.value))  # заполняем "строку таблицы из файла", каждую ячейку
        #        countColumns = countColumns + 1
        #    countColumns = 0
        #    print()

        # for numStr in range(1,sheet.max_row):
        #    rowPosition = self.ui.tableVar.rowCount()  # генерируем строку в таблице для записи в нее чиселок
        #    self.ui.tableVar.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"
        #    for numCol in range (1,sheet.max_column):
        #        self.ui.tableVar.setItem(rowPosition, numCol - 1, QtWidgets.QTableWidgetItem(sheet[numStr][numCol].value)) #  заполняем "строку таблицы из файла", каждую ячейку

        #    print(sheet[str][0].value)

        #    pageBook[str][0]:

        #    for i in range(1, sheet.max_row + 1)  # Начинаем c 1 строки, так как если начнем с 0 вылетит ошибка
        #       for j in range(0, sheet.max_column)  # Проходимся по кол-ву столбцов
        #           print(sheet[sheet[ i ][ j ].coordinate], end=" ")  # Выводим в консоль каждый элемент в каждой строке
        #       print()  # Перенос на новую строку

        #      Add text to the row
        #     for i in range (self.ui.tableVar.columnCount() - 1): # -1 потому что колонка "Прим." пустая
        #         self.ui.tableVar.setItem(rowPosition, i, QtWidgets.QTableWidgetItem(self.name)) #  заполняем "строку таблицы из файла", каждую ячейку
        #         self.ui.tableVar.setItem(rowPosition, i, QtWidgets.QTableWidgetItem(self.surname)) #
        #         self.ui.tableVar.setItem(rowPosition, i, QtWidgets.QTableWidgetItem(self.numGroup)) #             кроме "Прим."
        #         self.ui.tableVar.setItem(rowPosition, i, QtWidgets.QTableWidgetItem("CHECK")) #


if __name__ == "__main__":
    app = QApplication(sys.argv)

    MainWindow = WindowMenu()
    MainWindow1 = Window1()
    MainWindow2 = Window2()
    MainWindow3 = Window3()
    MainWindow4 = Window4()
    MainWindow5 = Window5()
    MainWindow6 = Window6()

    MainWindow.show()
    sys.exit(app.exec_())